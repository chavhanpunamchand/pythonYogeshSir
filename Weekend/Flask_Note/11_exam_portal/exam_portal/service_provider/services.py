from exam_portal.service_provider.config import app
from exam_portal.service_provider.model import *
from flask import request,session
import json
SUCCESS = "success"
ERROR = "error"


def send_email(username,password):
    pass

@app.route("/exam/portal/api/client/",methods=['POST'])
def client_registration():
    reqdata= request.get_json()         #Register(fullname='',email='') Login(username='',password='')
    if not reqdata:     # body check krtoy
        return json.dumps({ERROR:"Fields are mandatory [Fullname,email,username,password,role]"})
    else:
        fname = reqdata.get('fullname')     # each field
        user = reqdata.get('username')
        pwd = reqdata.get('password')
        email = reqdata.get('email')
        if not fname or not user or not pwd or not email:
            return json.dumps({ERROR: "Fields are mandatory [Fullname,email,username,password,role]"})

        if Login.query.filter_by(username=user).first():
            return json.dumps({ERROR: "Duplicate Username..!"})

        reg = Register(fullname=fname,email=email)
        db.session.add(reg)
        db.session.commit()

        login = Login(username=user,password=pwd,regid=reg.id,roleid=101)
        db.session.add(login)
        db.session.commit()
        #send_email(username,password)  -->
        return json.dumps({SUCCESS: f"{user} ADDED SUCCESSFULLY...!"})

def user_registration():
    pass

def upload_questions():
    pass

def delete_questions():
    pass

def update_questions():
    pass


from flask import render_template

def check_for_valid_session():
    try:
        if session['userinfo']:
            return True
    except:
        return False

@app.route('/service/mvc/role/',methods=["POST","GET"])  # do u really need to api -->
def add_role():
    if not check_for_valid_session():
        return   render_template('login.html', error='U need to login first to add role')

    if request.method == 'POST':
        reqdata = request.form
        if not reqdata or not reqdata.get('rolename') or not  reqdata.get('roleid'):
            return render_template('role.html', msg="All Fields Mandatory",
                               roles=Roles.query.filter(Roles.active == 'Y').all(),
                               role=Roles.get_dummy())
        else:
            roleid = reqdata.get('roleid')
            rolenm = reqdata.get('rolename')
            dbrole = Roles.query.filter_by(id=roleid).first()
            if dbrole:
                dbrole.role = rolenm
                db.session.commit()
                return  render_template('role.html', msg="Updated...",
                                       roles=Roles.query.filter(Roles.active == 'Y').all(),
                                       role=Roles(id=roleid,role=rolenm))

            else:
                role = Roles(id=roleid,role=rolenm)
                db.session.add(role)
                db.session.commit()
                return render_template('role.html', msg="Role ADDED Successfully...!",
                                   roles=Roles.query.filter(Roles.active == 'Y').all(),
                                   role=Roles.get_dummy())
    return render_template('role.html', msg="",
                           roles=Roles.query.filter(Roles.active == 'Y').all(),
                           role=Roles.get_dummy())

@app.route('/mvc/role/delete/<int:rid>',methods=['GET'])
def delete_role(rid):
    if not check_for_valid_session():
        return render_template('login.html', error='U need to login first to add role')

    role = Roles.query.filter_by(id=rid).first()
    if role:
        db.session.delete(role)
        db.session.commit()
        return render_template('role.html', msg = "Role Deleted Successfully...!",
                               roles = Roles.query.filter(Roles.active=='Y').all(),
                               role=Roles.get_dummy())


@app.route('/mvc/role/edit/<int:rid>',methods=['GET'])
def fetch_role_for_edit(rid):
    if not check_for_valid_session():
        return render_template('login.html', error='U need to login first to add role')

    return render_template('role.html', msg="",
                           roles=Roles.query.filter(Roles.active == 'Y').all(),
                           role=Roles.query.filter(Roles.id==rid,Roles.active=='Y').first())


ROLE_ID = "100"
def check_application_prerequisite():
    db.create_all()
    if not Roles.query.filter_by(id=ROLE_ID).first():

        role = Roles(id=100, role='ADMIN')
        db.session.add(role)
        db.session.commit()

        reg = Register(fullname='EXAMS', email='exams@gmail.com')
        db.session.add(reg)
        db.session.commit()

        login = Login(username="admin",password="admin123",regid=reg.id,roleid=ROLE_ID)
        db.session.add(login)
        db.session.commit()
        print('Admin User Created..')
    else:
        print('Admin is already created...')


@app.route('/exam/portal/login/',methods = ["GET","POST"])
def check_user_credentials():
    if request.method == 'POST':
        reqdata = request.form
        if not reqdata or not reqdata.get('username') or not reqdata.get('password'):
            return render_template('login.html',error='Username password required')
        else:
            login = Login.query.filter(Login.username==reqdata.get('username'),
                               Login.password==reqdata.get('password'),
                               Login.active=='Y').first()
            if login:
                session['userinfo'] = {'role':login.roleref.id,'username':login.username}
                return render_template('role.html', msg="",
                                       roles=Roles.query.filter(Roles.active == 'Y').all(),
                                       role=Roles.get_dummy())
            else:
                return render_template('login.html',error='Invalid Credentials')
    return render_template('login.html', error='')

@app.route('/exam/portal/logout/',methods = ['GET'])
def logout():
    msg = ''
    if check_for_valid_session():
       session.pop('userinfo')
       msg = 'Logged out..!'
    return render_template('login.html', error=msg)


def get_list_vendors():
    return Login.query.filter(Login.active=='Y',Login.roleid==101).all()

import openpyxl

FILE_PATH = 'D:\\python_work\\flask_projects\\exam_portal\\quessheets\\'




def read_excelsheet(filename,username):
    errors = {}
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook['questions']             # what if sheet name not present
    except BaseException as e:
        errors = {"file" : "{}".format(e.args)}
        print(e.args)
    else:
        max_row = sheet.max_row
        ques_list = []
        for row in range(1,max_row+1):
            if row==1:  # header
                continue
            a1 = sheet.cell(row, 1).value
            a2 = sheet.cell(row, 2).value
            a3 = sheet.cell(row, 3).value
            a4 = sheet.cell(row, 4).value
            a5 = sheet.cell(row,5).value
            a6 = sheet.cell(row,6).value                # if no 6 columns

            values = {a2,a3,a4,a5}
            if len(values) == 4 and a6 in values:
                ques_list.append(Questions(question=a1,option1=a2,option2=a3,option3=a4,option4=a5,answer=a6,username=username))
            else:
                errors[row]=(Questions(question=a1,option1=a2,option2=a3,option3=a4,option4=a5,answer=a6,username=username))

        if not errors:  # no errors -- tr ch question upload hotil
            db.session.add_all(ques_list)
            db.session.commit()

    return errors

#ques_list  = []
#allquestion = QUestion.query.all()  $ 18
#ranvals = set()
#while len(ranval)<5:
    #random.randint[0,len(allquestion)-1]
#for item in range(5)       # 5
        #ques.list(allquests[random.randint[0,len(allquestion)-1]])

@app.route("/upload/questions/",methods=['POST','GET'])
def upload_questions():
    msg = ''
    if request.method=='POST':
        username = request.form['user']     #request.get_json()         --> formdata--
        file = request.files['quesfile']
        print('username')
        print(file)
        file.save(FILE_PATH+username+'.xlsx')
        msg = 'Question Sheet uploaded..'

        errors = read_excelsheet(FILE_PATH+username+'.xlsx',username)
        if errors:
            msg = "Check Errors , {}".format(errors)
    return render_template('quesup.html',vendors = get_list_vendors(),msg=msg)

if __name__ == '__main__':
    check_application_prerequisite()
    app.run(debug=True)