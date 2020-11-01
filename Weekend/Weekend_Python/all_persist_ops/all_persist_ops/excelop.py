from all_persist_ops.service import EmpService
from all_persist_ops.employee import Employee
import openpyxl
workbook = None

FILENAME= 'employee.xlsx'



def get_workbook_instance():
    global workbook
    try:
        workbook = openpyxl.load_workbook(FILENAME) # if workbookpresent
        print('using existing new excel..!')
        #return workbook['empdata']          #returns sheet instance
        return workbook
    except FileNotFoundError as e:
        workbook = openpyxl.Workbook() # crate new
        sheet = workbook.create_sheet('empdata')
        #HEADERS
        sheet['A1'] = "EMP_ID"
        sheet['B1'] = "EMP_NAME"
        sheet['C1'] = "EMP_AGE"
        sheet['D1'] = "EMP_SALARY"
        sheet['E1'] = "EMP_ROLE"
        workbook.save(FILENAME)
        print('created new excel..!')
        return workbook


def remove_existingsheet_createheaders(workbook):
    workbook.remove(workbook['empdata'])  # existing data removed
    sheet = workbook.create_sheet('empdata')
    sheet['A1'] = "EMP_ID"
    sheet['B1'] = "EMP_NAME"
    sheet['C1'] = "EMP_AGE"
    sheet['D1'] = "EMP_SALARY"
    sheet['E1'] = "EMP_ROLE"
    workbook.save(FILENAME)
    return workbook

def close_workbook(wb):
    if wb:
        wb.close()
    else:
        print('Wb already closed..')


class EmpExcelServiceImpl(EmpService):
    def add_employee(self, emp):
        if type(emp) ==Employee:
            workbook = get_workbook_instance()
            sheet = workbook['empdata']
            noofrows = str(sheet.max_row+1)
            print(noofrows)
            sheet['A'+noofrows] = emp.empId
            sheet['B'+noofrows] = emp.empName
            sheet['C'+noofrows] = emp.empAge
            sheet['D'+noofrows] = emp.empSalary
            sheet['E'+noofrows] = emp.empRole
            workbook.save(FILENAME)
            print('Empdata {} written into excel file..!'.format(emp.empId))
        else:
            print('Invalid Emp info..!')

    def get_employee(self, empid):
        emps = self.get_all_employees()
        for emp in emps:
            if emp.empId==empid:
                return emp
        print('No emp with given id')

    def update_asper_user_choice(self):
        eid = int(input('Enter Id for update : '))

        options = {1: "Name",
                    2:"Age",
                    3:"Salary",
                    4: "Role"}

        if self.get_employee(eid):

            for key,value in options.items():
                print(str(key) +"."+str(value))

            choice = int(input('Enter Ur choice : '))
            if choice==1:
                value = input('Enter New Name : ')
            elif choice == 2:
                value = int(input('Enter New Age : '))
            elif choice == 3:
                value = float(input('Enter New Salary : '))
            elif choice == 4:
                value = input('Enter New Role : ')

            allemps = self.get_all_employees()
            for emp in allemps:
                if emp.empId == eid:
                    if choice==1:
                        emp.empName = value
                    elif choice==2:
                        emp.empAge = value
                    elif choice==3:
                        emp.empSalary = value
                    elif choice==4:
                        emp.empRole = value

            remove_existingsheet_createheaders(workbook)  # removed and created new headers
            for emp in allemps:  # 112 nasel
                self.add_employee(emp)  # newly created..
            print('Emp {} updated'.format(options.get(choice)))
        else:
            print("Emp with given id not exit..!")
    def get_all_employees(self):    # exe
        allemps = []
        workbook = get_workbook_instance()
        sheet = workbook['empdata']
        noofrows = sheet.max_row  # 4
        for item in range(noofrows):    # 0 1 2 3 --> indexing--> 0
            if item==0: #header asel    # 1 2 3 4 --> A1[header] B1
                continue
            item = item + 1       #1+1 -->2
            #print(item)
            eid = sheet['A' + str(item)].value  #A2
            enm = sheet['B' + str(item)].value
            eag = sheet['C' + str(item)].value
            esal = sheet['D' + str(item)].value
            erol = sheet['E' + str(item)].value
            allemps.append(Employee(eid,enm,eag,esal,erol))
        return allemps

    def delete_employee(self, empid):
        allemps = self.get_all_employees() # read entire sheet and get python list
        found=False  #
        for emp in allemps: #one by one row iterate-->
            if emp.empId == empid:  # id -->
                found = True         #row found
                allemps.remove(emp)  #remove --> memory--> not in excel
                print('Emp removed')

        if found: # true==?
            remove_existingsheet_createheaders(workbook)# removed and created new headers
            for emp in allemps: # 112 nasel
                self.add_employee(emp)  # newly created..
        else:
            print('Emp Id not found so cannot remove') # if not found

    def update_employee(self, empid, values):
        allemps = self.get_all_employees()
        found = False
        for emp in allemps:
            if emp.empId == empid:
                found = True
                emp.empName = values.empName
                emp.empSalary = values.empSalary
                emp.empAge = values.empAge
                emp.empRole = values.empRole
                print('Emp Updated')
                break
        if found:
            remove_existingsheet_createheaders(workbook)# removed and created new headers
            for emp in allemps:
                self.add_employee(emp)  # newly created..
        else:
            print('Emp Id not found so cannot Update') # if not found

    def emp_avg_salary(self):
        emplist = self.get_all_employees()  # will return file to python emp list
        avgsal = 0.0
        noofemps = len(emplist)  # no of emps in list
        totalSal = 0.0
        for emp in emplist:
            totalSal = totalSal + emp.empSalary
        avgsal = totalSal / noofemps
        print('Avg Salary -->', avgsal)
        return avgsal

    def get_rolebased_avg_salary(self, role):
        emplist = self.get_all_employees()
        noOfemps = 0
        totalSal = 0.0
        for emp in emplist:
            if emp.empRole == role:
                noOfemps = noOfemps + 1
                totalSal = totalSal + emp.empSalary
        avgsal = totalSal / noOfemps
        print('Manager\'s Avg Salary -->', avgsal)
        return avgsal

    def get_employees_based_on_role(self, rolenm):
        emplist = self.get_all_employees()
        newemplist = []
        for emp in emplist:
            if emp.empRole == rolenm:
                newemplist.append(emp)
        return newemplist

from all_persist_ops.helper import get_input_from_user
if __name__ == '__main__':
    #emp = get_input_from_user(idwant=True)
    excelService = EmpExcelServiceImpl()
    #excelService.add_employee(emp)
    print("Before ",excelService.get_all_employees())
    #print(excelService.get_employee(101))
    #print(excelService.emp_avg_salary())
    #eid = int(input('Enter Id for delete : '))
    #emp = get_input_from_user(idwant=False)
    #excelService.update_employee(eid,emp)
    #excelService.delete_employee(eid)
    #print("After -- ",excelService.get_all_employees())
    excelService.update_asper_user_choice()

    print("after ", excelService.get_all_employees())