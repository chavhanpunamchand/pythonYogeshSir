from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\Punamchand\Downloads\Python_Download\Python Classes\weekend_30-08-2020\weekend\OpenxlDwmo\Demofile.xlsx')

sheet = wb.active
sheet['A1'] = 'Ramesh Pawar'

sheet.cell(row=2, column=2).value = 5
wb.save(r'C:\Users\Punamchand\Downloads\Python_Download\Python Classes\weekend_30-08-2020\weekend\OpenxlDwmo\Demofile.xlsx')