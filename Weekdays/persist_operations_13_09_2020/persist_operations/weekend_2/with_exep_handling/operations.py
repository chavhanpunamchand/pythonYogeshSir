
from service import EmployeeService
from empinfo import Employee
#eid,enm,esal,eag,email

class FileNotPresent(BaseException):
    def __init__(self,msg):
        self.message =msg

def read_and_return_file_instance():
    try:
        file = open('emp.txt','r')
        return file             # return --> with block complete--> closed file
    except FileNotFoundError as e:
            raise FileNotPresent("Specific file not present.!")


def close_file_instance(file):
    try:
        if file:
            file.close()
    except:
        pass

class FileOperations(EmployeeService):

    def add_employee(self, emp):
        with open('emp.txt','a') as file:
            empinfo = str(emp.empId)+","+emp.empName +","+ str(emp.empSal) +","+ str(emp.empAge) +","+ emp.empEmail+"\n"
            file.writelines(empinfo)
            print('Emp Object Written into Text file..!')

    def get_employee(self, empid):
        fileinstance = None # file --        to access inside finally
        try:
            file = read_and_return_file_instance()  # risky code--> filenotfound        # excetion --. stacktrace--> clear
            fileinstance = file     # when this will be -> if no error -->
        except FileNotPresent as file:
            print(file.message)
        else:
            if file:
                all_lines = file.readlines()
                all_lines = [line.strip() for line in all_lines]
                for line in all_lines:
                    line = line.split(',')
                    try:
                        if empid == int(line[0]):
                            return Employee(line[0],line[1],line[2],line[3],line[4])
                    except IndexError as e:
                        print('Invalid Employee...!')
        finally:
              close_file_instance(fileinstance)



    def get_all_employees(self):
        all_emps = []
        with open('emp.txt', 'r') as file:
            all_lines = file.readlines()    #[l1,l2,l3,l4]
            all_lines = [line.strip() for line in all_lines]
            for empline in all_lines:
                line = empline.split(',')  #id,nm,age,sal,em
                all_emps.append(Employee(line[0], line[1], line[2], line[3], line[4]))
        return all_emps

    def remove_employee(self,eid):
        allemps  = self.get_all_employees()  #python -- emplist

        for emp in allemps: # first of all get -- object from list
            if emp.empId == eid:    # obj.empId -- eid
                allemps.remove(emp) # removed
                self.write_fresh_all_emps(allemps)
                return "{}, Emp removed".format(eid)


    def write_fresh_all_emps(self,allemps):
        with open('emp.txt','w') as file:       # this will remove existing file and writes new contents
            for emp in allemps: #
                empinfo = str(emp.empId) + "," + emp.empName + "," + str(emp.empSal) + "," + str(
                    emp.empAge) + "," + emp.empEmail + "\n"
                file.writelines(empinfo)

    def remove_all_employees(self):
        with open('emp.txt','w') as file:
            return []   # no emps

    def update_employee_info(self, empid, newval):
        allemps = self.get_all_employees()
        for emp in allemps:
            if emp.empId == empid:
                emp.empName = newval.empName
                emp.empEmail = newval.empEmail
                emp.empAge = newval.empAge
                emp.empSal = newval.empSal
                self.write_fresh_all_emps(allemps)
                print('Emp Record Updated..!')


import openpyxl


workinstance = None     #local -->  None

def get_old_or_new_oldbook(filename="empinfo"):
    global workinstance
    if workinstance:
        return workinstance
    else:
        try:
            workbook = openpyxl.load_workbook('{}.xlsx'.format(filename))
            workinstance = workbook
        except FileNotFoundError as e :     # incase not present
            workbook = openpyxl.Workbook()  # create new
            workbook.create_sheet('data')
            workbook.save('{}.xlsx'.format(filename))   #   save it
            workinstance = workbook
    return workinstance

get_old_or_new_oldbook() # init--> workinstance will be initi
from openpyxl.styles import PatternFill

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
flag = False
def create_headers(emp):
    global flag
    if not flag:
        sheet = workinstance['data']

        val = 65
        for name in emp.__dict__.keys(): # 5 veles -->
            sheet[chr(val)+'1'] = name.upper()
            sheet[chr(val) + '1'].fill = redFill
            val = val+1
            workinstance.save("empinfo.xlsx")
    flag=True

 #je operation --> almost --all --> python --> open
 #color --> font --> italic ---> openpyxl-->

class SalaryNotAsPerExpectations(BaseException):

    def __init__(self,msg,*args,**kwargs):
        self.message = msg

class ExcelSheetOperations:

    def add_employee(self, emp):    #   -->
        create_headers(emp)
        sheet = workinstance['Sheet']
        rowno = sheet.max_row+1   # +1
        # check all fields are present of not--> check-- fields as per business ahet or note
        if emp.empSal<20000.0:
            raise SalaryNotAsPerExpectations("Employee salary shud be atleast 20k")

        sheet['A{}'.format(rowno)] = emp.empId
        sheet['B{}'.format(rowno)] = emp.empName
        sheet['C{}'.format(rowno)] = emp.empAge
        sheet['D{}'.format(rowno)] = emp.empSal
        sheet['E{}'.format(rowno)] = emp.empEmail
        workinstance.save('empinfo.xlsx')
        return 'Emp record {} saved into excelsheet '.format(emp.empId)

    def get_employee(self, empid):
        pass

    def get_all_employees(self):
        sheet = workinstance['data']
        rows = sheet.max_row
        employees = []
        for enumr,item in enumerate(range(rows)): #100      # enum --> 0        --> row1
            if enumr==0:        # enum -- 0 -- row1
                continue  # skip as this header--->EmpID EmpName

            rownum = str(enumr+1)
            eid = sheet['A' + rownum].value
            enm = sheet['B' + rownum].value
            eag = sheet['C' + rownum].value
            esal = sheet['D' + rownum].value
            email = sheet['E' + rownum].value
            emp = Employee(eid, enm, esal, eag, email)
            employees.append(emp)
        return employees

    def remove_employee(self):
        pass

    def remove_all_employees(self):
        pass

    def update_employee_info(self, empid, newval):
        pass


# --> database operations --> selenium Automation-->    sat -->sun --> dev--> selenium -->

# weekdays --> ratri -->  8.30pm --> mon -->
#MRO --> Inheritance-->Database Operations -->
#
class DataBaseOperations:
    def add_employee(self, emp):
        pass

    def get_employee(self, empid):
        pass

    def get_all_employees(self):
        pass

    def remove_employee(self):
        pass

    def remove_all_employees(self):
        pass

    def update_employee_info(self, empid, newval):
        pass


class JSONFileOperations(EmployeeService):
    def add_employee(self, emp):
        pass

    def get_employee(self, empid):
        pass

    def get_all_employees(self):
        pass

    def remove_employee(self):
        pass

    def remove_all_employees(self):
        pass

    def update_employee_info(self, empid, newval):
        pass
